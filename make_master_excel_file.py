# -*- coding: utf-8 -*-
"""
Created on Thu Nov 10 10:16:54 2016

@author: pawel.cwiek
"""
# needs openpyxl 2.4 (and jdcal module is a dependency for it)

import openpyxl
import os
from time import strftime
import write_stats

import logging
#logging.basicConfig(level=logging.INFO)
log_path = os.path.join(os.getcwd(),'log.log')
print(log_path)
logg_config = dict([('level',logging.WARNING)])
logg_config.update(dict([('filename',log_path),('filemode','w')]))
logging.basicConfig(**logg_config)

def merge_data(master_wb_path, wb_paths, cells_range):
    """
        sums all the numeric fields from Excel files from wb_paths to form a Master Excel
        # if a numeric field is not to be summed (eg page number) - set it's field type to Text

            master_wb_path: string containig path to the copy of one of the other Excel files (to get the template/formatting) that will be used as master workbook

            workbook_paths: list containing strings for paths with Excel workbooks to merge

            cells_range: (row,col) tuple containig range of cells to be summed in each worsheet
    """
    def failure(file,sheet,cell):
        global error_count
        text = 'NOT summed due to wrong value in cell {}:{}{} in file: {}'.format(sheet,cell.column,cell.row,file.split(os.sep)[-1])
        cell.comment = openpyxl.comments.Comment(text,'Automatic tool')
        cell.comment.height = '200pt'
        cell.value = "N/A"
        error_count = error_count + 1

#    print(os.getcwd())
    global error_count
    error_count = 0
    print(master_wb_path, len(wb_paths), cells_range)
    mwb = openpyxl.load_workbook(master_wb_path, data_only = False)
    # clear data from master workbook to create a template
    # only fields that contain numbers(float,int) and are not locked
    # will not be zero'ed and will not have placeholder_text added as comment
    placeholder_text = 'field summed automatically from other files on: ' + strftime('%Y-%m-%d %H:%M:%S')
    analyzed_sheets = set()
    for sheet in mwb.get_sheet_names():
            msh = mwb.get_sheet_by_name(sheet)
            for rn in range(1, cells_range[0]):
                for cn in range(1,cells_range[1]):
                    mcell = msh.cell(row=rn, column=cn)

                    if mcell.protection.locked:
                        continue
                    analyzed_sheets.add(sheet)
#                    if mcell.value == None or mcell.data_type != 'n' :
#                        continue
#                    try:
#                        float(mcell.value)
#                    except ValueError:
#                        continue

#                    print(mcell.data_type)
#                    print(type(mcell.value))
#                    print(mcell.protection.locked)
                    mcell.value = 0
                    mcell.comment = openpyxl.comments.Comment(placeholder_text,'Automatic tool')
                    logging.info('mwb marked to write cell {}:{}{}'.format(sheet,openpyxl.utils.get_column_letter(cn),rn))

    for path in wb_paths:
        #readonly is False because if works faster (note no cwb.save is called so it's ok)
        cwb = openpyxl.load_workbook(filename = path, data_only=True, read_only=False)
        for sheet in analyzed_sheets:
            msh = mwb.get_sheet_by_name(sheet)
            csh = cwb.get_sheet_by_name(sheet)
            for rn in range(1, cells_range[0]+1):
                for cn in range(1,cells_range[1]+1):
                    mcell = msh.cell(row=rn, column=cn)
                    if mcell.protection.locked:
                        continue

                    try:
                        if mcell.comment == None:
                            continue
                    except AttributeError:
                        pass

                    if mcell.comment.text != placeholder_text:
                        continue
                    try:
                        ccell = csh.cell(row=rn, column=cn)
                    except IndexError:
                        logging.warning('''cell was not summed: could not load value from cell {}:{}{} from file: {}. Possibly end of table (range is to big)'''.format(sheet, openpyxl.utils.get_column_letter(cn), rn, path.split(os.path.sep)[-1]))
                        failure(path,sheet,mcell)
                        #raise
                        #continue
                        #continue

                    if ccell.value == None:
                        continue
                    logging.info('trying to sum cell {}:{}{} from {}'.format(sheet, openpyxl.utils.get_column_letter(cn), rn, path))
                    try:
                        mcell.value = mcell.value + ccell.value
                        logging.info('summed successfully')
                    except TypeError as err:
                        logging.warning('''Could not add data!
                        current fields:
                            worksheet= {}
                            cell= {}{}
                            masterwb value= {}
                            currentwb value= {}
                            current_file= {}
                        '''.format(sheet, openpyxl.utils.get_column_letter(cn), rn, mcell.value, ccell.value, path))
                        failure(path,sheet,mcell)

        cwb = None

    mwb.save(master_wb_path)
    try:
        if len(wb_paths) > 0: write_stats.dump_stats(wb_paths)
    except:
        pass
    return error_count

if __name__ == '__main__':
#    cwd = os.getcwd()
    cwd = r'C:\Users\pawel.cwiek\Desktop\New folder'
    os.chdir(cwd)

    mwb = """FCASTING LRa 16-11-15 Promenada ALL.xlsx"""
    workbook_paths = ["""FCASTING LRa 16-11-15 Promenada 01 PM.xlsx""","FCASTING LRa 16-11-15 Promenada 02 HVAC.xlsx","FCASTING LRa 16-11-15 Promenada 03 PH.xlsx","FCASTING LRa 16-11-15 Promenada 04 EL.xlsx"]
    cell_range = (300,100)

    merge_data(mwb,workbook_paths,cell_range)